import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def create_expert_xlsx(output_path):
    wb = openpyxl.Workbook()
    
    # --- Styles ---
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    sub_header_fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') # Light yellow for input
    calc_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid') # Light green for calc
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered = Alignment(horizontal='center', vertical='center')

    # --- Sheet 2: Parameters (基础资料) ---
    ws_param = wb.active
    ws_param.title = "基础资料"
    ws_param.append(["材料/参数名称", "数值", "单位", "备注"])
    params = [
        ["白板纸单价", 4500, "元/吨", "用于外管"],
        ["牛卡纸单价", 3800, "元/吨", "用于内管"],
        ["热熔胶单价", 0.8, "元/平方", ""],
        ["拉丝胶单价", 1.2, "元/平方", ""],
        ["正规纸规格(宽)", 787, "mm", ""],
        ["正规纸规格(长)", 1089, "mm", ""],
        ["大规纸规格(宽)", 889, "mm", ""],
        ["大规纸规格(长)", 1192, "mm", ""],
        ["正规纸单价", 0.9, "元/张", "157g铜版纸"],
        ["大规纸单价", 1.1, "元/张", "157g铜版纸"],
        ["纸箱每平方单价", 3.8, "元", "五层瓦楞纸箱"],
    ]
    for p in params:
        ws_param.append(p)
    
    # Formatting Parameters
    for cell in ws_param[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered
    ws_param.column_dimensions['A'].width = 25
    ws_param.column_dimensions['B'].width = 15

    # --- Sheet 3: LidDB (铁盖数据库) ---
    ws_lid = wb.create_sheet("铁盖数据库")
    ws_lid.append(["型号", "直径(mm)", "金银平底", "金银拉伸", "黑金平底", "黑金拉伸"])
    # Sample data based on previous extraction
    lids = [
        ["33", 33, 0.05, 0.05, 0.06, 0.06],
        ["40", 40, 0.06, 0.10, 0.07, 0.11],
        ["52", 52, 0.075, 0.10, 0.08, 0.12],
        ["65", 65, 0.105, 0.15, 0.11, 0.16],
        ["73", 73, 0.11, 0.17, 0.12, 0.18],
        ["83", 83, 0.13, 0.18, 0.15, 0.20],
        ["94", 94, 0.17, 0.23, 0.19, 0.25],
        ["99", 99, 0.17, 0.24, 0.19, 0.27],
        ["126", 126, 0.30, 0.36, 0.30, 0.39],
    ]
    for l in lids:
        ws_lid.append(l)

    # --- Sheet 1: Main Calculator (专家核价工作台) ---
    ws = wb.create_sheet("核价工作台", 0)
    
    # 1. Product Input
    ws.merge_cells('B2:E2')
    ws['B2'] = "一、产品基本参数录入"
    ws['B2'].font = Font(bold=True, size=12)
    
    input_labels = [
        ["产品直径(mm)", "B3", 94],
        ["产品高度(mm)", "B4", 150],
        ["订单数量(个)", "B5", 5000],
        ["内管厚度(mm)", "D3", 1.5],
        ["外管厚度(mm)", "D4", 1.0],
        ["目标毛利率(%)", "D5", 25]
    ]
    for label, pos, val in input_labels:
        ws[pos] = label
        val_pos = pos.replace('B', 'C').replace('D', 'E')
        ws[val_pos] = val
        ws[val_pos].fill = input_fill
        ws[val_pos].border = border

    # 2. Material Cost Calculation
    ws['B7'] = "二、材料成本计算 (自动)"
    ws['B7'].font = Font(bold=True)
    ws.append(["", "分项", "计算逻辑", "成本(元/个)"])
    
    # Formulas for Cost
    # Paper weight = Cir * Height * Thickness * Density
    # For simplicity, we keep the user's area logic but standardize it.
    # C3 is diameter. C4 is height.
    # Area (sqm) = (PI * Dia / 1000) * (Height / 1000)
    
    dia = "C3"
    height = "C4"
    qty = "C5"
    
    cost_rows = [
        ["外管成本", f"=PI()*{dia}/1000*{height}/1000 * 基础资料!B2 * 0.002", "D9"], # Using a simplified weight factor
        ["内管成本", f"=PI()*({dia}-2)/1000*({height}+10)/1000 * 基础资料!B3 * 0.002", "D10"],
        ["热熔胶", f"=PI()*{dia}/1000*{height}/1000 * 基础资料!B4", "D11"],
        ["盖子成本", "0.35", "D12"], # Placeholder for lookup
        ["面纸/印刷", "0.15", "D13"],
        ["纸箱分摊", "0.08", "D14"]
    ]
    
    curr_row = 9
    for name, logic, pos in cost_rows:
        ws.cell(row=curr_row, column=2, value=name)
        ws.cell(row=curr_row, column=3, value=logic)
        ws.cell(row=curr_row, column=4).fill = calc_fill
        ws.cell(row=curr_row, column=4).border = border
        curr_row += 1
    
    # 3. Labor & Overhead
    ws['B16'] = "三、人工与费用分摊"
    ws.append(["", "工序", "日工资(元)", "日产量(个)", "分摊(元/个)"])
    
    labor_data = [
        ["拉管/修头", 200, 5000],
        ["卷边/装配", 350, 6000],
        ["贴标", 200, 4000],
        ["厂租/水电", 1400, 15000]
    ]
    
    curr_row = 18
    for name, wage, t_qty in labor_data:
        ws.cell(row=curr_row, column=2, value=name)
        ws.cell(row=curr_row, column=3, value=wage).fill = input_fill
        ws.cell(row=curr_row, column=4, value=t_qty).fill = input_fill
        ws.cell(row=curr_row, column=5, value=f"=C{curr_row}/D{curr_row}").fill = calc_fill
        curr_row += 1
    
    # 4. Summary Output
    ws['B24'] = "四、核价结果汇总"
    ws['B24'].font = Font(bold=True, size=14, color='FF0000')
    
    summary_items = [
        ["总生产成本", "=SUM(D9:D14) + SUM(E18:E21)", "C25"],
        ["损耗(5%)", "=C25 * 0.05", "C26"],
        ["物流运费", "0.05", "C27"],
        ["最终总成本", "=SUM(C25:C27)", "C28"],
        ["建议含税售价", "=C28 / (1 - E5/100) * 1.13", "C30"],
        ["预计利润额", "=(C30/1.13) - C28", "C31"]
    ]
    
    for name, formula, pos in summary_items:
        ws[pos.replace('C', 'B')] = name
        ws[pos] = formula
        ws[pos].font = Font(bold=True)
        ws[pos].fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')

    # Data Validation for Lid lookup would be here in a real scenario
    
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15

    wb.save(output_path)
    print(f"Expert XLSX created at {output_path}")

output_path = "E:/123/常用文件/纸罐核价专家版.xlsx"
create_expert_xlsx(output_path)
