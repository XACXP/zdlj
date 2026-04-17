from openpyxl import load_workbook

file_path = 'E:/123/常用文件/纸筒计算试用版001(1).xlsx'
wb = load_workbook(file_path, data_only=True)
ws = wb['铁盖价格表']

print("--- 铁盖价格表 (Data) ---")
for row in ws.iter_rows(max_row=100, max_col=10, values_only=True):
    if any(row):
        print(row)
