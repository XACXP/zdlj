import openpyxl

def list_all():
    wb = openpyxl.load_workbook(r'E:/123/常用文件/纸筒计算试用版001(1) - 副本.xlsx')
    print(f"Sheets: {wb.sheetnames}")
    for sn in wb.sheetnames:
        sheet = wb[sn]
        print(f"--- Sheet: {sn} ---")
        for row in range(30, 35):
            r = [sheet.cell(row=row, column=c).value for c in range(1, 10)]
            print(f"Row {row}: {r}")

if __name__ == "__main__":
    list_all()
