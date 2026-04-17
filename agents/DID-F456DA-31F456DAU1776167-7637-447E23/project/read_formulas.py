from openpyxl import load_workbook

file_path = 'E:/123/常用文件/纸筒计算试用版001(1).xlsx'
wb = load_workbook(file_path, data_only=False) # Read formulas

sheets_to_read = ['全纸筒核价表', '铁盖纸筒核价表']

for name in sheets_to_read:
    print(f"\n--- {name} (Formulas) ---")
    ws = wb[name]
    for r_idx, row in enumerate(ws.iter_rows(max_row=30, max_col=20, values_only=True), 1):
        # Format output to show row number and cell values (formulas if present)
        print(f"Row {r_idx:2}: {row}")
