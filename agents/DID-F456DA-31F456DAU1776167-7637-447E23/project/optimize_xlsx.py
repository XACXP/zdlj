from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
import re

def clean_price(val):
    if val is None: return None
    if isinstance(val, (int, float)): return val
    s = str(val).replace('元', '').strip()
    try:
        return float(s)
    except:
        return val

def optimize_xlsx(input_path, output_path):
    wb = load_workbook(input_path)
    
    # 1. Clean up '铁盖价格表'
    if '铁盖价格表' in wb.sheetnames:
        ws = wb['铁盖价格表']
        # Columns C, E, H, J have prices
        for row in ws.iter_rows(min_row=4, max_row=40, min_col=3, max_col=10):
            for cell in row:
                if cell.column in [3, 5, 8, 10]:
                    cell.value = clean_price(cell.value)
        
        # Create a named range for the lookup table to make formulas cleaner
        # Range: B4:J40
        # Actually, it's better to just use direct references in this case.

    # 2. Optimize '铁盖纸筒核价表'
    if '铁盖纸筒核价表' in wb.sheetnames:
        ws = wb['铁盖纸筒核价表']
        
        # Formatting
        header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=5, max_row=5, min_col=2, max_col=15):
            for cell in row:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border

        # Add Data Validation for Diameter (B6)
        # Extract diameters from '铁盖价格表'
        lid_ws = wb['铁盖价格表']
        diameters = []
        for row in lid_ws.iter_rows(min_row=4, max_row=40, min_col=2, max_col=2, values_only=True):
            if row[0]:
                match = re.search(r'(\d+)', str(row[0]))
                if match:
                    diameters.append(match.group(1))
        
        if diameters:
            dv_dia = DataValidation(type="list", formula1=f'"{",".join(sorted(list(set(diameters)), key=int))}"', allow_blank=True)
            ws.add_data_validation(dv_dia)
            dv_dia.add(ws['B6'])
            # Set C6 to =B6/10
            ws['C6'] = "=B6/10"

        # Add Data Validation for Lid Type
        # We'll use a cell to select type. Let's use B7 for type selection.
        ws['B7'] = "金银平底" # Default
        dv_type = DataValidation(type="list", formula1='"金银平底,金银拉伸,黑金平底,黑金拉伸"', allow_blank=True)
        ws.add_data_validation(dv_type)
        dv_type.add(ws['B7'])
        ws['A7'] = "盖子类型:"
        ws['A7'].alignment = Alignment(horizontal='right')
        
        # Now the magic: Lookup the price in E11
        # B6 is diameter (e.g. 94)
        # B7 is type
        # In '铁盖价格表':
        # Column B matches diameter + "平底" or diameter + "黑底"
        # Column C is 金银平底, E is 金银拉伸, H is 黑金平底, J is 黑金拉伸
        # This is tricky because the labels in '铁盖价格表' are like "94平底" and "94黑底"
        
        # Formula for E11:
        # We can use VLOOKUP or a combination.
        # If B7 is "金银平底" or "金银拉伸", we look up B6 & "平底" in Col B, then return Col C or E.
        # If B7 is "黑金平底" or "黑金拉伸", we look up B6 & "黑底" in Col B, then return Col H or J (wait, Col H in '铁盖价格表' is "94黑底").
        
        formula = (
            '=IF(OR(B7="金银平底",B7="金银拉伸"), '
            'INDEX(IF(B7="金银平底", 铁盖价格表!$C$4:$C$40, 铁盖价格表!$E$4:$E$40), MATCH(B6&"平底", 铁盖价格表!$B$4:$B$40, 0)), '
            'INDEX(IF(B7="黑金平底", 铁盖价格表!$H$4:$H$40, 铁盖价格表!$J$4:$J$40), MATCH(B6&"黑底", 铁盖价格表!$G$4:$G$40, 0)))'
        )
        ws['E11'] = formula

    # 3. Final Polish for '全纸筒核价表'
    if '全纸筒核价表' in wb.sheetnames:
        ws = wb['全纸筒核价表']
        # Simple cleanup
        ws['C14'] = ws['C14'].value.replace('3.1415', 'PI()') if isinstance(ws['C14'].value, str) else ws['C14'].value
        ws['G14'] = ws['G14'].value.replace('3.1415', 'PI()') if isinstance(ws['G14'].value, str) else ws['G14'].value
        ws['C18'] = ws['C18'].value.replace('3.14159', 'PI()') if isinstance(ws['C18'].value, str) else ws['C18'].value
        ws['C17'] = "=C15*C14" # Ensure formulas are consistent

    wb.save(output_path)
    print(f"Optimized file saved to {output_path}")

input_file = 'E:/123/常用文件/纸筒计算试用版001(1).xlsx'
output_file = 'E:/123/常用文件/纸筒计算优化版.xlsx'
optimize_xlsx(input_file, output_file)
