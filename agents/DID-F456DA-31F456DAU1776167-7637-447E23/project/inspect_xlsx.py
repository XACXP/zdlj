from openpyxl import load_workbook

file_path = 'E:/123/常用文件/纸筒计算试用版001(1).xlsx'
wb = load_workbook(file_path, data_only=True)

print(f"Sheets: {wb.sheetnames}")

for name in wb.sheetnames:
    print(f"\n--- {name} ---")
    ws = wb[name]
    for row in ws.iter_rows(max_row=10, values_only=True):
        print(row)
