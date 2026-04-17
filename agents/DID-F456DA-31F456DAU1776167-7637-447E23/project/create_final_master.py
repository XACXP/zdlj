import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_final_expert_system(output_path):
    wb = openpyxl.Workbook()
    
    # --- Styles ---
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
    calc_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid')
    result_fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal='center', vertical='center')

    # --- Sheet 1: 基础参数库 (Master Data) ---
    ws_base = wb.active
    ws_base.title = "基础参数库"
    base_data = [
        ["分类", "参数名称", "数值", "单位", "备注"],
        ["厂房固定", "厂租 (年)", 250000, "元/年", "25万/年"],
        ["厂房固定", "水电 (月)", 20000, "元/月", "2万/月"],
        ["厂房固定", "标准月产量", 520000, "个/月", "日产2万*26天"],
        ["人力成本", "普通规格人工基准", 0.26, "元/个", "普通直径全工序"],
        ["人力成本", "16mm专机人工基准", 0.15, "元/个", "自动化专线"],
        ["人力成本", "拟购全能机人工基准", 0.08, "元/个", "48万全自动机"],
        ["原材料", "外管纸平方单价", 4.5, "元/㎡", ""],
        ["原材料", "内管纸平方单价", 4.0, "元/㎡", ""],
        ["原材料", "热熔胶平方单价", 0.8, "元/㎡", ""],
        ["纸张", "大规纸单价 (119*89)", 1.1, "元/张", ""],
        ["纸张", "正规纸单价 (108*78)", 0.9, "元/张", ""],
        ["工艺", "烫金单价 (平方)", 2.0, "元/㎡", ""],
        ["工艺", "烫金保底单价", 0.06, "元/个", "保底收费线"],
        ["工艺", "丝印UV (低墨量)", 0.08, "元/个", ""],
        ["工艺", "丝印UV (高墨量)", 0.12, "元/个", ""],
        ["工艺", "高端哑光覆膜加价", 0.14, "元/个", "针对精品版溢价"],
        ["物流", "默认国内运费分摊", 0.05, "元/个", "老外自提设为0"]
    ]
    for row in base_data: ws_base.append(row)
    
    # --- Sheet 2: 印刷报价表 (Printing) ---
    ws_print = wb.create_sheet("印刷报价表")
    print_data = [
        ["机型", "进纸规格 (mm)", "起印费 (元)", "万张单价 (元)", "备注"],
        ["8开机", "350*510", 220, 350, "适合小单"],
        ["4开机", "500*710", 350, 550, "适合大单"],
    ]
    for row in print_data: ws_print.append(row)

    # --- Sheet 3: 核价控制台 (Calculator) ---
    ws = wb.create_sheet("核价控制台", 0)
    
    # A. 输入区 (Inputs)
    ws.merge_cells('B2:F2')
    ws['B2'] = "东嘉纸罐核心报价工作台 V2.0 (自动化/大单专用版)"
    ws['B2'].font = Font(bold=True, size=14, color='2C3E50')
    
    labels = [
        ("产品直径 (mm)", "C4", 16), ("产品高度 (mm)", "E4", 112),
        ("订单数量 (个)", "C5", 1000000), ("生产模式", "E5", "专机自动化"),
        ("烫金面积 (c㎡)", "C6", 0), ("UV工艺要求", "E6", "无"),
        ("物流方式", "C7", "客户上门拉走"), ("目标毛利率 (%)", "E7", 25)
    ]
    for text, pos, val in labels:
        label_pos = pos.replace('C','B').replace('E','D')
        ws[label_pos] = text
        ws[pos] = val
        ws[pos].fill = input_fill
        ws[pos].border = border
    
    # Dropdowns
    dv_mode = DataValidation(type="list", formula1='"手工/半自动,专机自动化,新进全能机"', allow_blank=True)
    ws.add_data_validation(dv_mode); dv_mode.add(ws['E5'])
    dv_uv = DataValidation(type="list", formula1='"无,低墨量,高墨量"', allow_blank=True)
    ws.add_data_validation(dv_uv); dv_uv.add(ws['E6'])
    dv_ship = DataValidation(type="list", formula1='"送货上门,客户上门拉走"', allow_blank=True)
    ws.add_data_validation(dv_ship); dv_ship.add(ws['C7'])

    # B. 计算引擎 (Engine) - 隐藏计算过程
    # 材料费 (Materials)
    ws['B10'] = "一、原材料成本核算"
    ws['B10'].font = Font(bold=True)
    mat_rows = [
        ("纸管成本 (内外管+胶)", "=PI()*C4*(E4+20)/1000000 * (基础参数库!C8+基础参数库!C10)"),
        ("圆片成本 (上下2片)", "=((C4+10)*(C4+10)*2/1000000) * 基础参数库!C8"),
        ("面纸材料 (大规选优)", "=基础参数库!C11 / 80"), # 简化的开数，实际可更复杂
        ("原材料小计", "=SUM(D11:D13)")
    ]
    r = 11
    for name, formula in mat_rows:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=4, value=formula).fill = calc_fill
        r += 1

    # 工艺与印刷 (Process)
    ws['B16'] = "二、工艺与印刷分摊"
    ws['B16'].font = Font(bold=True)
    proc_rows = [
        ("印刷费分摊 (自动选机)", "=IF(C5>50000, MAX(印刷报价表!C3, (C5/4/10000)*印刷报价表!D3)/C5, MAX(印刷报价表!C2, (C5/2/10000)*印刷报价表!D2)/C5)"),
        ("烫金费用 (含保底)", "=MAX(基础参数库!C14, C6/10000 * 基础参数库!C13)"),
        ("丝印UV费用", "=IF(E6=\"无\", 0, IF(E6=\"低墨量\", 基础参数库!C15, 基础参数库!C16))"),
        ("工艺小计", "=SUM(D17:D19)")
    ]
    r = 17
    for name, formula in proc_rows:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=4, value=formula).fill = calc_fill
        r += 1

    # 固定分摊 (Fixed Overhead)
    ws['B21'] = "三、人工与厂房分摊"
    ws['B21'].font = Font(bold=True)
    overhead_rows = [
        ("直接人工成本", "=IF(E5=\"手工/半自动\", 基础参数库!C5, IF(E5=\"专机自动化\", 基础参数库!C6, 基础参数库!C7))"),
        ("厂租水电分摊", "=(基础参数库!C2/12 + 基础参数库!C3) / 基础参数库!C4"),
        ("管理/固定分摊小计", "=SUM(D22:D23)")
    ]
    r = 22
    for name, formula in overhead_rows:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=4, value=formula).fill = calc_fill
        r += 1

    # C. 汇总输出 (Summary)
    ws['B26'] = "四、报价汇总 (EXW/含税)"
    ws['B26'].font = Font(bold=True, size=12, color='FF0000')
    
    summary = [
        ("工厂总保本成本 (不含税)", "=(D14+D20+D24) * 1.05"), # 5% 损耗
        ("物流运费分摊", "=IF(C7=\"客户上门拉走\", 0, 基础参数库!C17)"),
        ("建议不含税单价 (目标利润)", "=(D27+D28) / (1 - E7/100)"),
        ("含税成交价 (13%税率)", "=D29 * 1.13")
    ]
    r = 27
    for name, formula in summary:
        ws.cell(row=r, column=2, value=name)
        ws.cell(row=r, column=4, value=formula).font = Font(bold=True)
        ws.cell(row=r, column=4).fill = result_fill
        ws.cell(row=r, column=4).border = border
        r += 1

    # Formatting columns
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20

    wb.save(output_path)
    print(f"Expert Master System created at {output_path}")

output_path = "E:/123/常用文件/东嘉纸罐专家核价系统_正式版.xlsx"
create_final_expert_system(output_path)
