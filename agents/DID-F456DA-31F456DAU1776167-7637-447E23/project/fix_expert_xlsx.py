import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

def create_fixed_expert_xlsx(output_path):
    wb = openpyxl.Workbook()
    
    # --- Styles ---
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid') 
    calc_fill = PatternFill(start_color='EBF1DE', end_color='EBF1DE', fill_type='solid') 
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    centered = Alignment(horizontal='center', vertical='center')

    # --- 基础资料 ---
    ws_param = wb.active
    ws_param.title = "基础资料"
    ws_param.append(["材料名称", "数值", "单位", "备注"])
    params = [
        ["外管纸单价", 3.5, "元/平方", "根据厚度和纸质定"],
        ["内管纸单价", 3.0, "元/平方", ""],
        ["热熔胶单价", 0.7, "元/平方", ""],
        ["正规纸单价", 0.9, "元/张", "157g铜版纸"],
        ["大规纸单价", 1.1, "元/张", "157g铜版纸"],
        ["铁盖单价", 0.35, "元/对", "默认参考值"],
        ["纸箱每平方单价", 3.8, "元", "五层瓦楞"],
    ]
    for p in params: ws_param.append(p)
    ws_param.column_dimensions['A'].width = 20

    # --- 核价工作台 ---
    ws = wb.create_sheet("核价工作台", 0)
    
    # 1. Input
    ws['B2'] = "一、产品基本参数"
    ws['B2'].font = Font(bold=True, size=12)
    inputs = [
        ("产品直径(mm)", "C3", 94), ("产品高度(mm)", "C4", 150),
        ("内管厚度(mm)", "E3", 1.5), ("订单数量(个)", "E4", 5000),
        ("目标毛利率(%)", "E5", 25)
    ]
    for label, pos, val in inputs:
        ws[pos.replace('C','B').replace('E','D')] = label
        ws[pos] = val
        ws[pos].fill = input_fill
        ws[pos].border = border

    # 2. Material Costs
    ws['B7'] = "二、原材料成本"
    ws['B7'].font = Font(bold=True)
    ws.append(["", "分项名称", "算法说明", "单项成本(元)"])
    
    # Dia: C3, Height: C4, Qty: E4
    mat_costs = [
        ["外管纸成本", "=PI()*C3*C4/1000000 * 基础资料!B2"],
        ["内管纸成本", "=PI()*(C3-E3*2)*C4/1000000 * 基础资料!B3"],
        ["热熔胶", "=PI()*C3*C4/1000000 * 基础资料!B4"],
        ["铁盖/配件", "基础资料!B7"],
        ["面纸费用", "基础资料!B5 / 10"], # Assume 10-up for placeholder
        ["纸箱分摊", "0.08"]
    ]
    
    row_idx = 9
    for name, formula in mat_costs:
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=formula.replace('=',''))
        ws.cell(row=row_idx, column=4, value=formula).fill = calc_fill
        ws.cell(row=row_idx, column=4).border = border
        row_idx += 1

    # 3. Labor Costs
    ws['B16'] = "三、加工与管理分摊"
    ws.append(["", "工序名称", "日工资(元)", "日产量(个)", "单价(元)"])
    labors = [
        ["拉管修头", 200, 5000], ["卷边装配", 350, 6000], 
        ["厂租水电", 1400, 15000], ["其他费用", 100, 10000]
    ]
    row_idx = 18
    for name, w, q in labors:
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=w).fill = input_fill
        ws.cell(row=row_idx, column=4, value=q).fill = input_fill
        ws.cell(row=row_idx, column=5, value=f"=C{row_idx}/D{row_idx}").fill = calc_fill
        row_idx += 1

    # 4. Total Result
    ws['B24'] = "四、核价结果汇总"
    ws['B24'].font = Font(bold=True, size=14, color='FF0000')
    summary = [
        ["原材料总计", "=SUM(D9:D14)"],
        ["加工费总计", "=SUM(E18:E21)"],
        ["总生产成本", "=B25 + B26"],
        ["含税建议售价", "=(B27+0.05)/(1-E5/100)*1.13"] # 0.05 is freight
    ]
    row_idx = 25
    for name, formula in summary:
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=formula).font = Font(bold=True)
        ws.cell(row=row_idx, column=3).fill = PatternFill(start_color='FDE9D9', end_color='FDE9D9', fill_type='solid')
        row_idx += 1

    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    wb.save(output_path)

output_path = "E:/123/常用文件/纸罐核价专家版_修正.xlsx"
create_fixed_expert_xlsx(output_path)
print("Updated fixed file.")
