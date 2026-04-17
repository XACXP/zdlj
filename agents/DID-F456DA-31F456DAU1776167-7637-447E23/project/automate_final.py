import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import re

def automate_paper_and_lid(input_path, output_path):
    wb = openpyxl.load_workbook(input_path)
    
    # 1. 准备铁盖数据源 (Prepare Lid Data Source)
    # 在“铁盖价格表”中建立直径列表
    lid_ws = wb['铁盖价格表']
    diameters = []
    # 遍历 B 列提取直径数字
    for row in lid_ws.iter_rows(min_row=4, max_row=40, min_col=2, max_col=2, values_only=True):
        if row[0]:
            match = re.search(r'(\d+)', str(row[0]))
            if match:
                diameters.append(match.group(1))
    
    unique_dias = sorted(list(set(diameters)), key=int)
    dia_list_str = ",".join(unique_dias)

    # 2. 优化“铁盖纸筒核价表” (Optimize Lid Tube Sheet)
    if '铁盖纸筒核价表' in wb.sheetnames:
        ws = wb['铁盖纸筒核价表']
        
        # A. 为直径 C6 添加下拉菜单 (Data Validation for Diameter)
        # 注意：原表 C6 是直径输入，将其设为下拉列表
        dv_dia = DataValidation(type="list", formula1=f'"{dia_list_str}"', allow_blank=True)
        ws.add_data_validation(dv_dia)
        dv_dia.add(ws['C6'])
        
        # B. 建立盖子类型选择 (Add Lid Type Selector)
        # 我们在 B7 放置一个选择框
        ws['B7'] = "金银平底" 
        dv_type = DataValidation(type="list", formula1='"金银平底,金银拉伸,黑金平底,黑金拉伸"', allow_blank=True)
        ws.add_data_validation(dv_type)
        dv_type.add(ws['B7'])
        ws['A7'] = "选择盖型:"
        ws['A7'].alignment = Alignment(horizontal='right')
        
        # C. 自动查价公式 (Auto Price Lookup Formula for Lid Cost)
        # 铁盖成本在 E11 (底盖) 或 E13 (拉深盖)
        # 我们统一用一个公式根据 B7 的选择来查
        lookup_formula = (
            '=IF(OR(B7="金银平底",B7="金银拉伸"), '
            'INDEX(IF(B7="金银平底", 铁盖价格表!$C$4:$C$40, 铁盖价格表!$E$4:$E$40), MATCH(C6&"平底", 铁盖价格表!$B$4:$B$40, 0)), '
            'INDEX(IF(B7="黑金平底", 铁盖价格表!$H$4:$H$40, 铁盖价格表!$J$4:$J$40), MATCH(C6&"黑底", 铁盖价格表!$G$4:$G$40, 0)))'
        )
        # 将结果放入 E11 (假设底盖位置)
        ws['E11'] = lookup_formula
        ws['E13'] = 0 # 拉深盖设为0，通过主公式选择

    # 3. 优化面纸费用联动 (Optimize Paper Cost Linkage)
    # 逻辑：将主表中的“面纸费用”单元格直接关联到“面纸开数”表的最优结果
    if '全纸筒核价表' in wb.sheetnames:
        ws_all = wb['全纸筒核价表']
        # 原表 G18 是纸张费用。我们将其关联到“面纸开数”表计算出的最优值
        # 根据之前的分析，最优价格公式在 面纸开数!K74 (正规) 或 M74 (大规)
        # 我们可以引用主表的“更划算是”单元格（假设是 S10）
        ws_all['G18'] = "=I9" # 引用计算器算出的费用
        
    # 4. 界面美化：高亮输入框 (Highlight Inputs)
    input_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    for name in ['全纸筒核价表', '铁盖纸筒核价表']:
        ws = wb[name]
        # 直径、高度、数量、膜厚、内管长等核心输入区
        for cell in ['C6', 'D6', 'E6', 'F6', 'G6', 'H6', 'I6']:
            ws[cell].fill = input_fill
        # 表头加固
        for cell in ws[5]:
            if cell.value:
                cell.fill = header_fill
                cell.font = header_font

    wb.save(output_path)
    print(f"Final Optimized file saved to {output_path}")

automate_paper_and_lid('E:/123/常用文件/纸筒计算改进中版001(1) .xlsx', 'E:/123/常用文件/纸筒计算全自动核价版.xlsx')
